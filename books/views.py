from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import load_workbook


def get_books(genre=None):
    workbook = load_workbook(filename='media/books.xlsx')
    worksheet = workbook.active

    books = []
    skip_first_row = True
    for row in worksheet.iter_rows(values_only=True, min_row=2 if skip_first_row else 1):
        genre_check = row[3].lower() == genre.lower() if genre else True
        if genre_check:
            book = {
               'book_id': int(row[0]),           
               'title': row[1],
               'author': row[2],
               'genre': row[3],
               'year': int(row[4]),
               'description': row[5],
               'cover': row[6]
           }
            books.append(book)
    return books
  
def genre_books(request, genre=None):
    books = get_books(genre)
    books_genre = sorted(set([x['genre'] for x in get_books()]))
    p = Paginator(books, 5)
    page_number = request.GET.get('page')
    page_obj = p.get_page(page_number)
    return render(request, 'genre_books.html', {'page_obj': page_obj,
                                                'books_genre': books_genre})

def book_detail(request, book_id):
    books = get_books()
    book = books[int(book_id) - 1] 
    return render(request, 'book_detail.html', {'book': book})

    